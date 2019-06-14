from django.db.models import *
from onlineApp.models import *
import  click
import  openpyxl
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE","classproject.settings")

import django
django.setup()

click.group()
def main():
    pass

@main.command()
def importCollegeData():
    wb = openpyxl.load_workbook("C:\\Users\\18suryateja\\Downloads\\students.xlsx")
    ws = wb['Colleges']
    for row in ws:
        temp = list(map(lambda x:x.value,row))
        if(temp[1]!='acronym'):
            c = College(name= temp[0],location =temp[2],acronym = temp[1],contact = temp[3])
            c.save()

@main.command()
def importStudentData():
    wb = openpyxl.load_workbook("C:\\Users\\18suryateja\\Downloads\\students.xlsx")
    ws = wb['Current']
    for row in ws:
        temp = list(map(lambda x:x.value,row))
        if(temp[0]!="Name"):
            s = Student(name = temp[0],email = temp[2],college = College.objects.get(acronym = temp[1]),db_folder = temp[3])
            s.save()

@main.command()
def importMockMarks():
    wb = openpyxl.load_workbook("./mock1_marks.xlsx")
    ws = wb['Sheet']
    for row in ws:
        temp = list(map(lambda x:x.value,row))
        if(temp[0]!=None and temp[0].strip()!='Student'):
            l=temp[0].split()
            name = l[2]
            m = MockTest1(student = Student.objects.filter(name = temp[0])[0],problem1=temp[1],problem2=temp[2],problem3=temp[3],problem4=temp[4],total=temp[5])
            m.save()

@main.command()
def getCollegesData():
    c = College.objects.all()
    for row in c:
        print("Name:",row.name,"Location:",row.location,"Acronym:",row.acronym,"Contact:",row.contact)

@main.command()
def getCollegeCount():
    return College.objects.count()

@main.command()
def getStudentsData():
    s = Student.objects.all()
    for row in s:
        print(row.name,row.college,row.dob,row.email,row.db_folder,row.dropped_out)

@main.command()
def getMockMarks():
    pass

main()